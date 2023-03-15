#Libreoffice imports 
import openpyxl
from openpyxl  import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border, Side
from openpyxl.styles import colors
from openpyxl.cell import cell

#RFID Scanner imports
import RPi.GPIO as GPIO
from mfrc522 import SimpleMFRC522
import time

#GUI import library
import tkinter as tk
import tkinter.font as tkFont

#Piezo buzzer set up 
GPIO.setmode(GPIO.BCM) 
buzzer = 23
GPIO.setup(buzzer,GPIO.OUT)

 

class App:
    def __init__(self, root):
        #setting title
        root.title("Point of Sales System")
        #setting window size
        width=380
        height=550
        screenwidth = root.winfo_screenwidth()
        screenheight = root.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height, (screenwidth - width) / 2, (screenheight - height) / 2)
        root.geometry(alignstr)
        root.resizable(width=False, height=False)

        GLabel_395=tk.Label(root)
        GLabel_395["bg"] = "#ff7800"
        ft = tkFont.Font(family='Times',size=10)
        GLabel_395["font"] = ft
        GLabel_395["fg"] = "#333333"
        GLabel_395["justify"] = "center"
        GLabel_395["text"] = "Item Description "
        GLabel_395.place(x=0,y=0,width=265,height=42)

        GLabel_476=tk.Label(root)
        GLabel_476["bg"] = "#ff7800"
        ft = tkFont.Font(family='Times',size=10)
        GLabel_476["font"] = ft
        GLabel_476["fg"] = "#333333"
        GLabel_476["justify"] = "center"
        GLabel_476["text"] = "Price"
        GLabel_476.place(x=260,y=0,width=118,height=42)

        GButton_339=tk.Button(root)
        GButton_339["activebackground"] = "#a3a3a3"
        GButton_339["activeforeground"] = "#000000"
        GButton_339["bg"] = "#9e1eff"
        ft = tkFont.Font(family='Times',size=10)
        GButton_339["font"] = ft
        GButton_339["fg"] = "#000000"
        GButton_339["justify"] = "center"
        GButton_339["text"] = "Complete Transaction "
        GButton_339.place(x=0,y=520,width=125,height=30)
        GButton_339["command"] = self.GButton_339_command

        GLabel_6=tk.Label(root)
        GLabel_6["bg"] = "#a61fff"
        GLabel_6["disabledforeground"] = "#393d49"
        ft = tkFont.Font(family='Times',size=10)
        GLabel_6["font"] = ft
        GLabel_6["fg"] = "#333333"
        GLabel_6["justify"] = "center"
        GLabel_6["text"] = "Total"
        GLabel_6.place(x=0,y=440,width=70,height=25)

    def GButton_339_command(self):
        print("command")




wb = Workbook()
sheet = wb['Sheet']


title = sheet.cell(1,3,'Point of Sales System')
subtitle1 = sheet.cell(2,1,'Item Code')
subtitle2 = sheet.cell(2,2,'UID')
subtitle3 = sheet.cell(2,3,'Item Details')
subtitle4 = sheet.cell(2,4,'Item Price')
subtitle5 = sheet.cell(2,5,'Qty')
subtitle6 = sheet.cell(2,6,'Barcode ID')
subtitle7 = sheet.cell(18,4,'Total:')

fill1 = PatternFill(patternType = 'solid',start_color= '33AFFF')
fill2 = PatternFill(patternType = 'solid',start_color = 'BFC9CA')

sheet['C1'].fill = fill1
sheet['D1'].fill = fill1
sheet['A2'].fill = fill2
sheet['B2'].fill = fill2
sheet['C2'].fill = fill2
sheet['D2'].fill = fill2
sheet['E2'].fill = fill2
sheet['F2'].fill = fill2
sheet['D18'].fill = fill2

GPIO.setwarnings(False) #Get rid of GPIO overflow warning 


reader = SimpleMFRC522()
total = 0
count = 0
tru = 1
i = 3
j = 6

xiguadata = ["xigua",5]
bananadata = ["banana",3]
receipt = [bananadata,bananadata,bananadata]

cart = []

#Function to add an item to the cart
def add_item(item_name, item_price, item_id):
    item = {"name": item_name, "price": item_price, "id":item_id}
    cart.append(item)
    
#Fucntion to print out the shopping cart
def print_cart():
    print("Shopping Cart:")
    for item in cart:
        print("- ", item["name"], ": $", item["price"])

#Function to calculate price in total
def calculate_total():
    total = 0
    for item in cart:
        total += ["price"]
    return total


try:
    while tru:
        
        id = 0
        id,text = reader.read() #Pulls from the Write2.py file 
        print(text) #Write in the name given to items
        #print(id) #Write in the UID
        GPIO.output(buzzer,GPIO.HIGH) #Piezo Buzzer to turn sound each time a RFID is scanned 
        time.sleep(2) # Delay the time between each scan
        
        #Add to the receipt
        if(id == 483755500000): #BANANA
            add_item("banana", 3, 483755500000)
            total = total + 3
            receipt[count] = bananadata
            count = count + 1
            item1 = sheet.cell(i,3,"banana")
            cost1 = sheet.cell(i,4,"$3.00")
            subtitle7 = sheet.cell(i,2,id)
            subtitle8 = sheet.cell(i,2,id)
            subtitle9 = sheet.cell(i,2,id)
            i = i+1
        
        if(id == 483990839803): #XIGUA
            add_item("xigua", 5, 483990839803)
            total = total + 5
            receipt[count] = xiguadata
            count = count + 1
            item2 = sheet.cell(i,3,"xigua")
            cost2 = sheet.cell(i,4,"$5.00")
            subtitle7 = sheet.cell(i,2,id)
            subtitle8 = sheet.cell(i,2,id)
            subtitle9 = sheet.cell(i,2,id)
            i = i+1
        
        print_cart()  
        
        #After 3 added items, go ahead and print receipt.
        if(count == 3):
            print("Here's your receipt!")
        if(count == 3):
            for row in receipt:
                for column in row:
                    print(column, end=' ')
                print()
        if(count == 3):
            tru = 0
            print("Your total is:",total)
            subtitle7 = sheet.cell(18,5, total)
           
finally:
    GPIO.cleanup()



wb.save('POS_System.xls')

if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()






